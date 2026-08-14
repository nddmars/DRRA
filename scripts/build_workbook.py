#!/usr/bin/env python3
"""
Build the WSG experiment-data workbook from the DRRA harness output.

Reads results/paper_metrics.json and results/feedback.json (running the
experiments to produce them if they are missing) and writes
results/WSG_Experiment_Data.xlsx. The Defensibility Index cells are live Excel
formulas referencing a Parameters sheet, so the workbook recalculates when any
input or weight is edited.

Usage:
    python scripts/build_workbook.py
"""

from __future__ import annotations

import importlib.util
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RESULTS = os.path.join(REPO, "results")
OUT = os.path.join(RESULTS, "WSG_Experiment_Data.xlsx")

FONT = "Arial"
ACCENT = "1F4E79"
HDR_FILL = PatternFill("solid", fgColor=ACCENT)
SUB_FILL = PatternFill("solid", fgColor="DCE6F1")
BAND = PatternFill("solid", fgColor="F2F2F2")
INPUT_BLUE = Font(name=FONT, color="0000FF", size=10)
FORMULA_BLACK = Font(name=FONT, color="000000", size=10)
BOLD = Font(name=FONT, bold=True, size=10)
HDRF = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE = Font(name=FONT, bold=True, size=14, color=ACCENT)
NOTE = Font(name=FONT, italic=True, size=9, color="595959")
GREEN = Font(name=FONT, bold=True, color="375623", size=10)
_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _load(name, rel):
    spec = importlib.util.spec_from_file_location(name, os.path.join(REPO, rel))
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod


def _ensure_results():
    """Load cached results, generating them if absent."""
    pm = os.path.join(RESULTS, "paper_metrics.json")
    fb = os.path.join(RESULTS, "feedback.json")
    if not os.path.exists(pm):
        exp = _load("wsg_experiment", "scripts/run_experiment.py")
        os.makedirs(RESULTS, exist_ok=True)
        json.dump(exp.run_experiment(30), open(pm, "w"), indent=2)
    if not os.path.exists(fb):
        fbm = _load("wsg_feedback", "scripts/run_feedback_experiment.py")
        os.makedirs(RESULTS, exist_ok=True)
        json.dump(fbm.run(10), open(fb, "w"), indent=2)
    return json.load(open(pm)), json.load(open(fb))


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDRF
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _box(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


# DI parameter cell references (Parameters sheet)
T_DRRT, SLA, EPS = "Parameters!$B$4", "Parameters!$B$5", "Parameters!$B$6"
WA, WB_, WC, WD = "Parameters!$B$7", "Parameters!$B$8", "Parameters!$B$9", "Parameters!$B$10"


def _di_formula(mttd, mttc, apcr, rf):
    D = f"MAX({EPS},MIN(1,1-{mttd}/{T_DRRT}))"
    C = f"MAX({EPS},MIN(1,1-{mttc}/{SLA}))"
    P = f"MAX({EPS},MIN(1,1-{apcr}))"
    R = f"MAX({EPS},MIN(1,{rf}))"
    denom = f"({WA}/({D})+{WB_}/({C})+{WC}/({P})+{WD}/({R}))"
    return f"=({WA}+{WB_}+{WC}+{WD})/{denom}"


def build(paper, fb) -> Workbook:
    wb = Workbook()

    # --- Parameters ---
    ws = wb.active
    ws.title = "Parameters"
    ws["A1"] = "Defensibility Index — Parameters"
    ws["A1"].font = TITLE
    rows = [
        ("Parameter", "Value", "Notes"),
        ("T_drrt (detection deadline, s)", 300, "Threat-response deadline for MTTD normalization"),
        ("MTTC SLA (s)", 90, "SHIELD containment SLA"),
        ("epsilon", 0.000001, "Floor to keep components in (0,1] for the harmonic mean"),
        ("Weight alpha — detection", 0.30, "MTTD efficiency"),
        ("Weight beta — containment", 0.30, "MTTC efficiency"),
        ("Weight gamma — prevention", 0.25, "1 - APCR"),
        ("Weight delta — recovery", 0.15, "Recovery fidelity"),
    ]
    for i, (a, b, c) in enumerate(rows):
        r = 3 + i
        ws.cell(r, 1, a); ws.cell(r, 2, b); ws.cell(r, 3, c)
        if i == 0:
            _style_header(ws, r, 3)
        else:
            ws.cell(r, 1).font = BOLD
            ws.cell(r, 2).font = INPUT_BLUE
            ws.cell(r, 3).font = NOTE
            for cc in (1, 2, 3):
                ws.cell(r, cc).border = BORDER
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 52
    ws["A12"] = "Blue = input values. DI formulas on the other sheets reference these cells."
    ws["A12"].font = NOTE

    # --- Table 4 ---
    ws = wb.create_sheet("Table 4 - Measured")
    ws["A1"] = "Table 4 — Measured WSG Simulation Results"
    ws["A1"].font = TITLE
    ws["A2"] = ("Source: DRRA harness  python scripts/run_experiment.py --reps 30  "
                "(IsolationForest). Mean ± 95% CI over 30 repetitions.")
    ws["A2"].font = NOTE
    hdr = ["Metric", "Scenario A (Change Healthcare)", "± 95% CI",
           "Scenario B (MOVEit)", "± 95% CI", "Baseline (No WSG)"]
    hr = 4
    for c, h in enumerate(hdr, 1):
        ws.cell(hr, c, h)
    _style_header(ws, hr, len(hdr))
    A = paper["conditions"]["A_change_healthcare"]
    B = paper["conditions"]["B_moveit"]
    data = [
        ("VIGIL MTTD (s)", A["mttd_seconds"]["mean"], A["mttd_seconds"]["ci95"],
         B["mttd_seconds"]["mean"], B["mttd_seconds"]["ci95"], "N/A"),
        ("SHIELD MTTC (s)", A["mttc_seconds"]["mean"], A["mttc_seconds"]["ci95"],
         B["mttc_seconds"]["mean"], B["mttc_seconds"]["ci95"], "N/A"),
        ("False Positive Rate", A["false_positive_rate"], "", B["false_positive_rate"], "", "—"),
        ("Attack Path Completion Rate (APCR)", A["apcr"]["mean"], A["apcr"]["ci95"],
         B["apcr"]["mean"], B["apcr"]["ci95"], 1.0),
        ("Recovery Fidelity", A["recovery_fidelity"]["mean"], A["recovery_fidelity"]["ci95"],
         B["recovery_fidelity"]["mean"], B["recovery_fidelity"]["ci95"], 1.0),
    ]
    r = hr + 1
    for name, am, ac, bm, bc, base in data:
        ws.cell(r, 1, name).font = BOLD
        for col, val in [(2, am), (3, ac), (4, bm), (5, bc), (6, base)]:
            cell = ws.cell(r, col, val)
            cell.font = INPUT_BLUE if isinstance(val, (int, float)) else FORMULA_BLACK
            cell.alignment = CENTER
        if (r - hr) % 2 == 0:
            for c in range(1, 7):
                ws.cell(r, c).fill = BAND
        r += 1
    for rr in (hr + 3, hr + 4, hr + 5):
        for col in (2, 4, 6):
            cell = ws.cell(rr, col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
    for rr in (hr + 1, hr + 2):
        for col in (2, 3, 4, 5):
            ws.cell(rr, col).number_format = "0.00"
    di_row = r
    ws.cell(di_row, 1, "Defensibility Index (DI)").font = BOLD
    ws.cell(di_row, 2, _di_formula(f"B{hr+1}", f"B{hr+2}", f"B{hr+4}", f"B{hr+5}"))
    ws.cell(di_row, 4, _di_formula(f"D{hr+1}", f"D{hr+2}", f"D{hr+4}", f"D{hr+5}"))
    ws.cell(di_row, 6, _di_formula("300", "90", f"F{hr+4}", f"F{hr+5}"))
    for col in (2, 4, 6):
        ws.cell(di_row, col).number_format = "0.000"
        ws.cell(di_row, col).alignment = CENTER
        ws.cell(di_row, col).font = GREEN
    ws.cell(di_row, 6).comment = Comment(
        "Baseline: APCR=100% drives prevention to ~0; the harmonic mean collapses DI to ~0.", "DRRA")
    _box(ws, hr, 1, di_row, 6)
    ws.cell(di_row + 2, 1, "DI is computed live from the metric rows via the Parameters weights.").font = NOTE
    ws.column_dimensions["A"].width = 34
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A5"

    # --- Table 6 ---
    ws = wb.create_sheet("Table 6 - Comparative")
    ws["A1"] = "Table 6 — Comparative Defensibility Index"
    ws["A1"].font = TITLE
    ws["A2"] = ("WSG row: measured (Table 4 mean). Comparator rows: stated capability assumptions "
                "(NOT measured). DI computed live via the formula.")
    ws["A2"].font = NOTE
    hdr = ["Architecture", "MTTD (s)", "MTTC (s)", "Recovery Fidelity", "False Positive Rate",
           "APCR", "Prevention (1-APCR)", "Defensibility Index"]
    hr = 4
    for c, h in enumerate(hdr, 1):
        ws.cell(hr, c, h)
    _style_header(ws, hr, len(hdr))
    comp = [
        ("Conventional ML detection", 12.0, 900.0, 0.930, 0.048, 0.90),
        ("SOAR-based automation", 15.0, 45.0, 0.945, 0.039, 0.55),
        ("Backup-centric recovery", 300.0, 1800.0, 0.975, 0.045, 0.98),
        ("Proposed WSG (measured)", 2.5, 7.8, 1.000, 0.000, 0.46),
    ]
    r = hr + 1
    for name, mttd, mttc, rf, fpr, apcr in comp:
        ws.cell(r, 1, name).font = BOLD if "WSG" in name else FORMULA_BLACK
        for col, val, pct in [(2, mttd, False), (3, mttc, False), (4, rf, True),
                              (5, fpr, True), (6, apcr, True)]:
            cell = ws.cell(r, col, val)
            cell.font = INPUT_BLUE
            cell.alignment = CENTER
            cell.number_format = "0.0%" if pct else "0.0"
        ws.cell(r, 7, f"=1-F{r}")
        ws.cell(r, 7).number_format = "0.0%"
        ws.cell(r, 7).font = FORMULA_BLACK
        ws.cell(r, 7).alignment = CENTER
        ws.cell(r, 8, _di_formula(f"B{r}", f"C{r}", f"F{r}", f"D{r}"))
        ws.cell(r, 8).number_format = "0.000"
        ws.cell(r, 8).font = GREEN
        ws.cell(r, 8).alignment = CENTER
        if "WSG" in name:
            for c in range(1, 9):
                ws.cell(r, c).fill = SUB_FILL
        r += 1
    _box(ws, hr, 1, r - 1, 8)
    note = ("Comparator assumptions: Conventional ML = detection but manual containment (~15 min); "
            "SOAR = automated containment within SLA, weaker prevention; "
            "Backup-centric = strong recovery, no detection/containment.")
    ws.cell(r + 1, 1, note).font = NOTE
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=8)
    ws.cell(r + 1, 1).alignment = LEFT
    ws.column_dimensions["A"].width = 28
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 15
    ws.freeze_panes = "A5"

    # --- Feedback ---
    ws = wb.create_sheet("Feedback (Sec 5.5)")
    ws["A1"] = "Section 5.5 — Closed-Loop Feedback Validation"
    ws["A1"].font = TITLE
    ws["A2"] = ("Source: python scripts/run_feedback_experiment.py --cycles 10. "
                "Two-stage detector; benign batches fed back as confirmed-benign labels each cycle.")
    ws["A2"].font = NOTE
    hdr = ["Cycle", "Primary-only FPR", "Ensemble FPR", "Detection Rate", "Defensibility Index"]
    hr = 4
    for c, h in enumerate(hdr, 1):
        ws.cell(hr, c, h)
    _style_header(ws, hr, len(hdr))
    r = hr + 1
    for x in fb["per_cycle"]:
        ws.cell(r, 1, x["cycle"]).font = INPUT_BLUE
        for col, key, pct in [(2, "primary_fpr", True), (3, "ensemble_fpr", True),
                              (4, "detection_rate", True), (5, "defensibility_index", False)]:
            cell = ws.cell(r, col, x[key])
            cell.font = INPUT_BLUE
            cell.number_format = "0.0%" if pct else "0.000"
            cell.alignment = CENTER
        if (r - hr) % 2 == 0:
            for c in range(1, 6):
                ws.cell(r, c).fill = BAND
        r += 1
    _box(ws, hr, 1, r - 1, 5)
    last = r - 1
    sr = r + 1
    ws.cell(sr, 1, "Summary").font = BOLD
    summ = [
        ("Initial ensemble FPR", f"=C{hr+1}", "0.0%"),
        ("Final ensemble FPR", f"=C{last}", "0.0%"),
        ("Relative FPR reduction", f"=IF(C{hr+1}=0,0,(C{hr+1}-C{last})/C{hr+1})", "0.0%"),
        ("Initial DI", f"=E{hr+1}", "0.000"),
        ("Final DI", f"=E{last}", "0.000"),
        ("DI improvement", f"=E{last}-E{hr+1}", "0.000"),
    ]
    for i, (lab, f, fmt) in enumerate(summ):
        rr = sr + 1 + i
        ws.cell(rr, 1, lab).font = BOLD
        cell = ws.cell(rr, 2, f)
        cell.number_format = fmt
        cell.font = FORMULA_BLACK
    _box(ws, sr, 1, sr + len(summ), 2)
    ws.column_dimensions["A"].width = 24
    for col in "BCDE":
        ws.column_dimensions[col].width = 17
    ws.freeze_panes = "A5"

    # --- README ---
    ws = wb.create_sheet("README")
    ws["A1"] = "WSG / DRRA — Experiment Data Workbook"
    ws["A1"].font = TITLE
    lines = [
        "",
        "Measured results from the DRRA reference implementation (github.com/nddmars/DRRA).",
        "Every Defensibility Index cell is a live formula that recalculates from the metric",
        "inputs and the weights on the Parameters sheet.",
        "",
        "Sheets:",
        "  Parameters — DI weights and thresholds (edit the blue cells to re-tune).",
        "  Table 4 - Measured — WSG results, 30 reps, mean +/- 95% CI (Scenarios A/B + baseline).",
        "  Table 6 - Comparative — WSG (measured) vs three architecture classes (assumptions).",
        "  Feedback (Sec 5.5) — false-positive reduction and DI across 10 feedback cycles.",
        "",
        "Colour key:  BLUE = input/measured    BLACK = formula    GREEN = Defensibility Index",
        "",
        "Reproduce:",
        "  python scripts/run_experiment.py --reps 30",
        "  python scripts/run_feedback_experiment.py --cycles 10",
        "  python scripts/build_workbook.py",
    ]
    for i, ln in enumerate(lines):
        ws.cell(2 + i, 1, ln).font = NOTE if ln.startswith("  ") else Font(name=FONT, size=10)
    ws.column_dimensions["A"].width = 95

    wb.move_sheet("README", -(len(wb.sheetnames) - 1))
    return wb


def main():
    paper, fb = _ensure_results()
    wb = build(paper, fb)
    os.makedirs(RESULTS, exist_ok=True)
    wb.save(OUT)
    print("wrote", os.path.relpath(OUT, REPO))


if __name__ == "__main__":
    main()
